"""Phase 1 Definition of Done 검증 — 실제 HTTP 요청으로 전체 흐름을 확인한다."""

import datetime as dt

from sqlalchemy import select

from app import models
from tests.conftest import app_session, login_as


def _create_retreat(client, name="2026 여름수련회 Belong", cap=8000):
    response = client.post(
        "/retreats/create",
        data={
            "name": name,
            "start_date": "2026-07-20",
            "end_date": "2026-07-23",
            "meal_subsidy_per_person": cap,
            "clone_from": "",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200, response.text
    with app_session() as db:
        return db.scalars(select(models.Retreat).where(models.Retreat.name == name)).one()


def _create_departments(client, names):
    for name in names:
        response = client.post(
            "/departments/create", data={"name": name, "color_tag": ""}, follow_redirects=True
        )
        assert response.status_code == 200
    with app_session() as db:
        return list(db.scalars(select(models.Department).order_by(models.Department.id)))


def _create_categories(client, rows):
    for l1, l2, l3, amount in rows:
        response = client.post(
            "/budget/categories",
            data={"level1": l1, "level2": l2, "level3": l3, "planned_amount": amount},
            follow_redirects=True,
        )
        assert response.status_code == 200
    with app_session() as db:
        return list(db.scalars(select(models.BudgetCategory).order_by(models.BudgetCategory.id)))


def _dim_by_task_title(html: str) -> dict[str, bool]:
    """할 일 카드마다 '제목 → 흐리게(dim) 표시 여부'를 뽑아낸다."""
    import re

    result = {}
    for match in re.finditer(
        r'<div class="item ([^"]*)">.*?<div class="item-title">(.*?)</div>', html, re.S
    ):
        title = re.sub(r"<[^>]+>", "", match.group(2)).strip()
        if title:
            result[title] = "faded" in match.group(1)
    return result


DEFAULT_CATEGORIES = [
    ("홍보", "포스터", "인쇄비", 300_000),
    ("시스템", "음향", "렌탈", 800_000),
    ("장소비", "숙소", "", 4_000_000),
    ("식비", "본행사 식사", "자율배식", 2_500_000),
    ("그 외", "수련회 준비지원", "모임 식사비", 1_000_000),
]


# ------------------------------------------------------------------ 로그인/권한


def test_로그인하지_않으면_로그인_화면으로_보낸다(client):
    response = client.get("/", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_최초_로그인_사용자는_총무팀_관리자가_된다(client):
    login_as(client, "01011112222", name="총무 김간사")

    with app_session() as db:
        user = db.scalars(select(models.User)).one()
    assert user.role == "admin"
    assert user.name == "총무 김간사"


def test_로그인_화면은_초대_링크로_들어오라고만_말한다(client):
    """비밀번호도 인증번호도 없다 (CLAUDE.md 4-12)."""
    response = client.get("/login")

    assert response.status_code == 200
    assert "초대 링크" in response.text
    # 인증번호를 입력받는 자리가 없어야 한다 (문구에 그 낱말이 나오는 것과는 다르다)
    assert "data-dev-code" not in response.text
    assert 'name="code"' not in response.text
    assert 'name="phone_number"' not in response.text


def test_아무_링크나_넣으면_들어올_수_없다(client):
    response = client.get("/invite/아무거나-지어낸-토큰", follow_redirects=False)

    assert response.status_code == 403
    assert "링크를 찾을 수 없습니다" in response.text
    with app_session() as db:
        assert db.scalars(select(models.User)).all() == []


# ------------------------------------------------------ DoD: 회차·부서·예산 카테고리


def test_총무팀은_회차와_부서3개_예산항목5개를_등록할_수_있다(admin_client):
    _create_retreat(admin_client)
    departments = _create_departments(admin_client, ["총무팀", "홍보팀", "찬양팀"])
    categories = _create_categories(admin_client, DEFAULT_CATEGORIES)

    assert len(departments) == 3
    assert len(categories) == 5

    page = admin_client.get("/budget")
    assert page.status_code == 200
    assert "포스터" in page.text
    assert "8,600,000" in page.text  # 예산 합계


def test_회차를_복제하면_부서와_예산항목이_따라온다(admin_client):
    source = _create_retreat(admin_client)
    _create_departments(admin_client, ["총무팀", "홍보팀", "찬양팀"])
    _create_categories(admin_client, DEFAULT_CATEGORIES)

    admin_client.post(
        "/retreats/create",
        data={
            "name": "2027 겨울수련회",
            "start_date": "",
            "end_date": "",
            "meal_subsidy_per_person": 6000,
            "clone_from": str(source.id),
        },
        follow_redirects=True,
    )

    with app_session() as db:
        new = db.scalars(
            select(models.Retreat).where(models.Retreat.name == "2027 겨울수련회")
        ).one()
        depts = db.scalars(
            select(models.Department).where(models.Department.retreat_id == new.id)
        ).all()
        cats = db.scalars(
            select(models.BudgetCategory).where(models.BudgetCategory.retreat_id == new.id)
        ).all()

    assert [d.name for d in depts] == ["총무팀", "홍보팀", "찬양팀"]
    assert len(cats) == 5
    assert new.meal_subsidy_per_person == 6000


# ------------------------------------------------------ DoD: 부서별 Task 흐리게 표시


def _setup_two_departments(admin_client):
    _create_retreat(admin_client)
    depts = _create_departments(admin_client, ["홍보팀", "찬양팀"])
    return depts


def test_부서리더는_자기_부서_할일만_선명하게_보고_타부서는_흐리게_본다(admin_client, client):
    hongbo, chanyang = _setup_two_departments(admin_client)

    admin_client.post(
        "/tasks/create",
        data={"title": "포스터 시안 확정", "department_id": str(hongbo.id), "status": "대기"},
        follow_redirects=True,
    )
    admin_client.post(
        "/tasks/create",
        data={"title": "콘티 정리", "department_id": str(chanyang.id), "status": "대기"},
        follow_redirects=True,
    )
    admin_client.post(
        "/users/create",
        data={
            "name": "홍보 리더",
            "phone_number": "010-3333-4444",
            "role": "dept_lead",
            "department_id": str(hongbo.id),
        },
        follow_redirects=True,
    )

    leader = client
    login_as(leader, "01033334444")
    page = leader.get("/tasks?scope=all")

    assert page.status_code == 200
    # 타 부서 할 일도 화면에 존재하되 흐리게(dim) 표시된다
    assert "포스터 시안 확정" in page.text
    assert "콘티 정리" in page.text

    dimmed = _dim_by_task_title(page.text)
    assert dimmed["포스터 시안 확정"] is False  # 내 부서 → 선명하게
    assert dimmed["콘티 정리"] is True  # 타 부서 → 흐리게


def test_부서리더는_타부서_할일의_상태를_바꿀_수_없다(admin_client, client):
    hongbo, chanyang = _setup_two_departments(admin_client)
    admin_client.post(
        "/tasks/create",
        data={"title": "콘티 정리", "department_id": str(chanyang.id), "status": "대기"},
        follow_redirects=True,
    )
    admin_client.post(
        "/users/create",
        data={
            "name": "홍보 리더",
            "phone_number": "010-3333-4444",
            "role": "dept_lead",
            "department_id": str(hongbo.id),
        },
        follow_redirects=True,
    )
    with app_session() as db:
        task = db.scalars(select(models.Task)).one()

    login_as(client, "01033334444")
    response = client.post(f"/tasks/{task.id}/status", data={"status": "완료"})

    assert response.status_code == 403
    with app_session() as db:
        assert db.get(models.Task, task.id).status == "대기"


def test_열람전용_계정은_할일을_만들_수_없다(admin_client, client):
    hongbo, _ = _setup_two_departments(admin_client)
    admin_client.post(
        "/users/create",
        data={
            "name": "담당 전도사",
            "phone_number": "010-5555-6666",
            "role": "viewer",
            "department_id": str(hongbo.id),
        },
        follow_redirects=True,
    )

    login_as(client, "01055556666")
    response = client.post(
        "/tasks/create", data={"title": "몰래 추가", "department_id": str(hongbo.id)}
    )

    assert response.status_code == 403
    with app_session() as db:
        assert db.scalars(select(models.Task)).all() == []


# ------------------------------------------------------------ DoD: 지출·예산 진행률


def test_지출을_등록하면_예산_진행률에_바로_반영된다(admin_client):
    _create_retreat(admin_client)
    categories = _create_categories(admin_client, DEFAULT_CATEGORIES)
    poster = categories[0]  # 예산 300,000

    response = admin_client.post(
        "/expenses/create",
        data={
            "budget_category_id": str(poster.id),
            "expense_date": dt.date.today().isoformat(),
            "amount": "100000",
            "payer_name": "김총무",
            "payer_account": "국민 123-456",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    budget_page = admin_client.get("/budget")
    assert "33.3" in budget_page.text  # 100,000 / 300,000
    assert "200,000" in budget_page.text  # 잔액

    dashboard = admin_client.get("/dashboard")
    assert "100,000" in dashboard.text


def test_식대_지출은_인원수만_넣으면_지원금액과_개인부담이_자동_계산된다(admin_client):
    _create_retreat(admin_client, cap=8000)
    categories = _create_categories(admin_client, DEFAULT_CATEGORIES)
    meal_cat = categories[4]

    response = admin_client.post(
        "/expenses/create",
        data={
            "budget_category_id": str(meal_cat.id),
            "expense_date": dt.date.today().isoformat(),
            "amount": "130600",
            "is_meal_expense": "1",
            "meal_headcount": "12",
            "meal_attendees": "박민준 홍성헌 민주아",
            "level3b": "모임 식사비-1",
            "payer_name": "박민준",
            "payer_account": "국민 123-456",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    with app_session() as db:
        entry = db.scalars(select(models.ExpenseEntry)).one()

    assert entry.subsidy_amount == 96_000  # min(130,600, 12 × 8,000)
    assert entry.personal_burden_amount == 34_600
    assert entry.meal_attendee_names == ["박민준", "홍성헌", "민주아"]
    assert entry.receipt_number == 1

    page = admin_client.get("/expenses")
    assert "96,000" in page.text
    assert "34,600" in page.text


def test_식대_상한은_회차_설정을_따른다(admin_client):
    retreat = _create_retreat(admin_client, cap=5000)
    admin_client.post(
        "/expenses/create",
        data={"amount": "100000", "is_meal_expense": "1", "meal_headcount": "10"},
        follow_redirects=True,
    )

    with app_session() as db:
        entry = db.scalars(select(models.ExpenseEntry)).one()

    assert retreat.meal_subsidy_per_person == 5000
    assert entry.subsidy_amount == 50_000
    assert entry.personal_burden_amount == 50_000


def test_식대는_지원금액만_예산에서_집행된_것으로_집계된다(admin_client):
    _create_retreat(admin_client, cap=8000)
    categories = _create_categories(admin_client, DEFAULT_CATEGORIES)

    admin_client.post(
        "/expenses/create",
        data={
            "budget_category_id": str(categories[4].id),
            "amount": "130600",
            "is_meal_expense": "1",
            "meal_headcount": "12",
        },
        follow_redirects=True,
    )

    page = admin_client.get("/budget")
    assert "96,000" in page.text
    assert "130,600" not in page.text


def test_환급_대상자_목록에_지출자별_합계가_나온다(admin_client):
    _create_retreat(admin_client, cap=8000)
    for amount, head in [("68900", "9"), ("130600", "12")]:
        admin_client.post(
            "/expenses/create",
            data={
                "amount": amount,
                "is_meal_expense": "1",
                "meal_headcount": head,
                "payer_name": "박민준",
                "payer_account": "국민 123456-01-123456",
            },
            follow_redirects=True,
        )

    page = admin_client.get("/refunds")

    assert "박민준" in page.text
    assert "국민 123456-01-123456" in page.text
    assert "164,900" in page.text  # 68,900 + 96,000


def test_지급여부를_전환할_수_있다(admin_client):
    _create_retreat(admin_client)
    admin_client.post("/expenses/create", data={"amount": "50000"}, follow_redirects=True)
    with app_session() as db:
        entry = db.scalars(select(models.ExpenseEntry)).one()
    assert entry.paid is False

    admin_client.post(
        f"/expenses/{entry.id}/paid", data={"redirect_to": "/expenses"}, follow_redirects=True
    )

    with app_session() as db:
        updated = db.get(models.ExpenseEntry, entry.id)
    assert updated.paid is True
    assert updated.paid_date == dt.date.today()


# ------------------------------------------------------------------ DoD: 엑셀 export


def test_지출_내역을_엑셀로_내려받을_수_있다(admin_client):
    _create_retreat(admin_client)
    categories = _create_categories(admin_client, DEFAULT_CATEGORIES)
    admin_client.post(
        "/expenses/create",
        data={
            "budget_category_id": str(categories[4].id),
            "amount": "130600",
            "is_meal_expense": "1",
            "meal_headcount": "12",
            "meal_attendees": "박민준 홍성헌",
            "payer_name": "박민준",
        },
        follow_redirects=True,
    )

    response = admin_client.get("/export/expenses.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["지출 상세내역", "예산 대비 집행", "환급 대상자"]

    ws = wb["지출 상세내역"]
    header = [c.value for c in ws[1]]
    assert header[:2] == ["구분", "항목"]
    row = [c.value for c in ws[2]]
    assert 130600 in row and 96000 in row and 34600 in row

    refunds = wb["환급 대상자"]
    assert refunds["A2"].value == "박민준"
    assert refunds["C2"].value == 96000


# ------------------------------------------------------------------ 일정


def test_일자와_일정을_등록하고_Day별_탭으로_볼_수_있다(admin_client):
    _create_retreat(admin_client)
    admin_client.post(
        "/schedule/days", data={"label": "1일차", "date": "2026-07-20"}, follow_redirects=True
    )
    admin_client.post(
        "/schedule/days", data={"label": "2일차", "date": "2026-07-21"}, follow_redirects=True
    )
    with app_session() as db:
        day = db.scalars(select(models.ScheduleDay).order_by(models.ScheduleDay.id)).first()

    admin_client.post(
        f"/schedule/days/{day.id}/items",
        data={
            "title": "1일차 집회",
            "start_time": "19:30",
            "end_time": "21:30",
            "location": "대강당",
        },
        follow_redirects=True,
    )

    page = admin_client.get(f"/schedule?day_id={day.id}")

    assert "1일차" in page.text and "2일차" in page.text
    assert "1일차 집회" in page.text
    assert "19:30" in page.text
    assert "대강당" in page.text


def test_회차가_없으면_안내_화면을_보여준다(admin_client):
    page = admin_client.get("/")

    assert page.status_code == 200
    assert "수련회 회차가 없습니다" in page.text


# ------------------------------------------------------------------ 영수증 업로드


def test_영수증_파일을_첨부해서_지출을_등록할_수_있다(admin_client):
    _create_retreat(admin_client)
    png = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01"
        b"\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
    )

    response = admin_client.post(
        "/expenses/create",
        data={"amount": "12000", "payer_name": "김총무"},
        files={"receipt": ("receipt.png", png, "image/png")},
        follow_redirects=True,
    )
    assert response.status_code == 200

    with app_session() as db:
        entry = db.scalars(select(models.ExpenseEntry)).one()
    assert entry.receipt_file_url is not None
    assert entry.receipt_file_url.startswith("/uploads/")

    stored = admin_client.get(entry.receipt_file_url)
    assert stored.status_code == 200
    assert stored.content == png


def test_허용되지_않는_확장자는_거부한다(admin_client):
    _create_retreat(admin_client)

    response = admin_client.post(
        "/expenses/create",
        data={"amount": "12000"},
        files={"receipt": ("악성.exe", b"MZ", "application/octet-stream")},
    )

    assert response.status_code == 400
    with app_session() as db:
        assert db.scalars(select(models.ExpenseEntry)).all() == []


def test_로그인하지_않으면_영수증_파일을_볼_수_없다(admin_client, client):
    _create_retreat(admin_client)
    admin_client.post(
        "/expenses/create",
        data={"amount": "12000"},
        files={"receipt": ("receipt.png", b"\x89PNG\r\n\x1a\n", "image/png")},
        follow_redirects=True,
    )
    with app_session() as db:
        url = db.scalars(select(models.ExpenseEntry)).one().receipt_file_url

    response = client.get(url, follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_경로_조작으로_다른_파일을_읽을_수_없다(admin_client):
    response = admin_client.get("/uploads/..%2F..%2Fapp.db")

    assert response.status_code in (400, 404)


# ------------------------------------------------------------------ 열람 전용


def test_열람전용_계정은_지출을_등록할_수_없다(admin_client, client):
    _create_retreat(admin_client)
    admin_client.post(
        "/users/create",
        data={
            "name": "담당 전도사",
            "phone_number": "010-5555-6666",
            "role": "viewer",
            "department_id": "",
        },
        follow_redirects=True,
    )

    login_as(client, "01055556666")
    response = client.post("/expenses/create", data={"amount": "99999"})

    assert response.status_code == 403
    with app_session() as db:
        assert db.scalars(select(models.ExpenseEntry)).all() == []


def test_열람전용_계정도_화면은_모두_볼_수_있다(admin_client, client):
    _create_retreat(admin_client)
    _create_categories(admin_client, DEFAULT_CATEGORIES)
    admin_client.post(
        "/users/create",
        data={"name": "담당 전도사", "phone_number": "010-5555-6666", "role": "viewer"},
        follow_redirects=True,
    )

    login_as(client, "01055556666")

    for path in ["/", "/schedule", "/tasks", "/budget", "/expenses", "/refunds", "/settings"]:
        assert client.get(path).status_code == 200, path


# ------------------------------------------------------------------ 오류 화면


def test_없는_주소로_들어가면_안내_화면을_보여준다(admin_client):
    response = admin_client.get("/이런-주소는-없습니다", headers={"accept": "text/html"})

    assert response.status_code == 404
    assert "홈으로" in response.text  # JSON 이 아니라 사람이 읽는 화면


def test_없는_회차를_열면_안내_화면을_보여준다(admin_client):
    response = admin_client.get("/schedule?retreat_id=99999", headers={"accept": "text/html"})

    assert response.status_code == 404
    assert "회차를 찾을 수 없습니다" in response.text
