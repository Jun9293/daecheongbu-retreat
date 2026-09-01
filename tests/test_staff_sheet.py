"""봉사자 시간표 (CLAUDE.md 5-8). 수용 기준 1~16.

**표 구조를 만드는 함수를 직접 시험한다.** 화면과 파일을 각각 시험하면 둘이
어긋나는 것을 못 잡는다 — 어긋남이 이 기능에서 가장 비싼 실패다.
12번은 둘이 같은 구조에서 나왔는지를 본다.

시각에 기대는 것이 없으므로 now 를 주입할 자리는 '내려받은 시각' 하나뿐이다.
"""

from __future__ import annotations

import datetime as dt
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from app import models
from app.domain import staff_sheet, staff_xlsx
from tests.conftest import app_session, login_as

OPEN = dt.date(2026, 8, 21)
CLOSE = dt.date(2026, 8, 23)


@pytest.fixture
def sheet_data(admin_client):
    """실제 자료의 모양을 그대로 줄인 것 — 선발대는 전체일정이 없고,
    2일차에는 자정 넘긴 프로그램 둘이 같은 시각에 있고, 나란히 도는 것도 있다."""
    with app_session() as db:
        retreat = models.Retreat(
            name="2026 여름수련회 Belong", start_date=OPEN, end_date=CLOSE)
        db.add(retreat)
        db.flush()
        for order, (key, name) in enumerate(
            [("chongmuM", "1 총무M"), ("hebron", "5 헤브론"), ("koram", "6 코람데오")]
        ):
            db.add(models.Department(
                retreat_id=retreat.id, key=key, name=name,
                color_tag="#888888", sort_order=order))

        def program(day, time, name, *, audience="staff", track="main",
                    parallel=False, end=None, items=(), host=None, place=None):
            p = models.Program(
                retreat_id=retreat.id, day=day, start_time=time, name=name,
                host=host, place=place, audience=audience, track=track,
                parallel=parallel, end_time=end, sort_order=0)
            db.add(p)
            db.flush()
            for order, (part, who, text) in enumerate(items):
                db.add(models.ProgramItem(
                    program_id=p.id, phase="pre", part_key=part,
                    assignee_name=who, text=text, sort_order=order,
                    scope="team" if part in ("헤브론", "코람데오") else "person"))
            return p

        # 선발대 — 전체일정이 하나도 없다
        program("선발대", "10:00", "짐정리", items=[
            ("비품", "서윤", "총무팀 개인 항목"),           # 표에 들어가면 안 된다
            ("코람데오", "재하", "짐 나르기"),              # 프로그램과 같은 칸 (섞임)
        ])
        program("선발대", "22:00", "영수증 정리", track="ops")   # ops — 안 들어간다

        # 1일차 — 전체일정 있음 + 나란히 도는 것
        program("1일차", "10:00", "봉사자 예배", audience="all", host="한지훈 목사",
                items=[("헤브론", "건우", "음향 장비 확인")])
        program("1일차", "19:00", "집회1", audience="all")
        program("1일차", "19:00", "새친구", audience="all",
                parallel=True, end="23:00")
        program("1일차", "14:00", "리허설")
        program("1일차", "16:00", "무대 점검", audience="all", items=[
            ("코람데오", "재하", "무대 세팅"),
        ])

        # 2일차 — **자정 넘긴 시각 둘이 같은 줄에서 시작한다**
        program("2일차", "00:00", "광고", audience="all")
        program("2일차", "00:00", "야식", audience="all")
        program("2일차", "08:00", "아침식사", audience="all", host="총무팀")

        # ── 아래는 실제 자료에서 그대로 가져온 것 ──
        # 프로그램과 그 프로그램에 붙은 봉사팀 항목이 **같은 말**인 경우들
        program("2일차", "08:00", "골든벨 준비", host="담당M · 헤브론 · 총무팀",
                items=[("헤브론", "담당M·헤브론·총무팀", "골든벨 준비")])
        program("선발대", "14:00", "음향 및 무대설치", host="헤브론",
                items=[("헤브론", "헤브론", "음향 및 무대설치 (3.5h)")])
        program("선발대", "18:00", "홀 튜닝", host="헤브론",
                items=[("헤브론", "헤브론", "홀 튜닝 (1.5h)")])
        program("폐회", "16:00", "본당 도착 · 하차", host="전체",
                items=[("헤브론", "헤브론", "하차")])
        # 담당이 사람 이름이고 장소가 있는 것
        program("폐회", "10:30", "폐회 예배 · 성찬", audience="all",
                host="한지훈 목사 · 최도현M", place="강당")
        db.commit()
        return {"retreat_id": retreat.id}


def build(db, data):
    return staff_sheet.build(db, db.get(models.Retreat, data["retreat_id"]))


def cell_at(sheet, row, col):
    for cell in sheet["cells"]:
        if (cell["row"] <= row < cell["row"] + cell["rowspan"]
                and cell["col"] <= col < cell["col"] + cell["colspan"]):
            return cell
    return None


def texts(sheet):
    return " | ".join(c["text"] for c in sheet["cells"] if c.get("text"))


# ---------------------------------------------------------------- 1·2. 화면


def test_01_봉사팀_보기_탭이_있고_누구나_본다(admin_client, sheet_data, client):
    page = admin_client.get("/live/staff")
    assert page.status_code == 200
    assert "봉사팀 보기" in page.text
    assert "엑셀로 내려받기" in page.text

    # 진행 화면에도 건너가는 자리가 있다
    assert "/live/staff" in admin_client.get("/live?stay=1").text

    # 로그인하지 않으면 못 본다
    assert client.get("/live/staff", follow_redirects=False).status_code in (303, 401)


def test_01b_부서원도_본다(sheet_data, admin_client):
    """봉사팀도 총무팀도 본다 — 총무팀을 거치지 않아야 한다."""
    with app_session() as db:
        dept = db.scalars(select(models.Department).where(
            models.Department.key == "hebron")).first()
        db.add(models.User(name="헤브론 팀원", phone_number="01055556666",
                           role="member", department_id=dept.id))
        db.commit()

    from app.main import app

    member = TestClient(app)
    login_as(member, "01055556666")
    assert member.get("/live/staff").status_code == 200
    assert member.get("/live/staff.xlsx").status_code == 200


def test_02_봉사팀_소속이면_그_탭이_기본으로_열린다(sheet_data, admin_client):
    with app_session() as db:
        for key, phone, name in [("hebron", "01055556666", "헤브론 리더"),
                                 ("koram", "01055557777", "코람데오 리더")]:
            dept = db.scalars(select(models.Department).where(
                models.Department.key == key)).first()
            db.add(models.User(name=name, phone_number=phone,
                               role="dept_lead", department_id=dept.id))
        db.commit()

    from app.main import app

    for phone in ("01055556666", "01055557777"):
        who = TestClient(app)
        login_as(who, phone)
        landing = who.get("/live", follow_redirects=False)
        assert landing.status_code == 303
        assert landing.headers["location"] == "/live/staff"

    # 총무팀은 진행 화면이 기본이다
    assert admin_client.get("/live", follow_redirects=False).status_code == 200


# ---------------------------------------------------------------- 3. 축


def test_03_시각_열이_양쪽_끝에_있고_한_시각이_두_줄이다(sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)

    left, right = sheet["time_cols"]
    assert left == 0
    assert right == sheet["total_cols"] - 1
    assert sheet["columns"][left]["kind"] == "time"
    assert sheet["columns"][right]["kind"] == "time"

    # 06:00 부터 다음날 01:00 까지, 한 시각이 두 줄
    assert staff_sheet.BODY_ROWS == (25 - 6 + 1) * 2
    assert staff_sheet.row_label(0) == "06:00"
    assert staff_sheet.row_label(1) is None          # 30분 줄에는 안 붙는다
    assert staff_sheet.row_label(2) == "07:00"
    assert staff_sheet.row_label(staff_sheet.BODY_ROWS - 2) == "01:00"

    # 시각 칸은 두 줄씩 병합돼 있다
    first = cell_at(sheet, sheet["header_rows"], 0)
    assert first["rowspan"] == 2 and first["text"] == "06:00"


# ---------------------------------------------------------------- 4. 빈 칸


def test_04_전체일정이_없는_날은_그_칸_자체가_없다(sheet_data):
    """빈 칸을 두면 "왜 비었지" 를 묻게 된다 — 선발대가 그렇다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    by_day = {d["name"]: d["bands"] for d in sheet["days"]}
    assert by_day["선발대"] == ["staff"]
    assert by_day["1일차"] == ["staff", "all"]
    assert by_day["2일차"] == ["staff", "all"]

    # 봉사자 칸이 먼저, 전체일정 칸이 나중
    day1 = [c for c in sheet["columns"] if c.get("day") == "1일차"]
    assert [c["band"] for c in day1] == ["staff", "staff", "all", "all"]


# ---------------------------------------------------------------- 5. 걸러낸 것


def test_05_ops_프로그램과_총무팀_개인_항목이_표에_없다(sheet_data):
    """이것이 이 표의 이유다 — 봉사팀이 보는 것은 총무팀의 실행 목록이 아니다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    body = texts(sheet)
    assert "영수증 정리" not in body        # ops 프로그램
    assert "총무팀 개인 항목" not in body    # 총무팀 파트 항목

    # 봉사팀 파트 항목은 들어간다
    assert "음향 장비 확인" in body
    assert "무대 세팅" in body
    # 봉사자·전체일정 프로그램도 들어간다
    assert "짐정리" in body and "봉사자 예배" in body


# ---------------------------------------------------------------- 6. parallel


def test_06_parallel_은_오른쪽_열에_서고_왼쪽이_반_칸이_된다(sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)

    day1_all = next(c for c in sheet["columns"]
                    if c.get("day") == "1일차" and c["band"] == "all")
    base = sheet["columns"].index(day1_all)

    집회 = next(c for c in sheet["cells"] if c["text"].startswith("집회1"))
    새친구 = next(c for c in sheet["cells"] if c["text"].startswith("새친구"))

    assert 집회["col"] == base and 집회["colspan"] == 1, "왼쪽이 반 칸이 아니다"
    assert 새친구["col"] == base + 1, "나란히 도는 것이 오른쪽 열이 아니다"
    assert 집회["row"] == 새친구["row"]

    # end_time 이 있으면 거기까지 (19:00~23:00 = 8줄)
    assert 새친구["rowspan"] == 8

    # 나란히 도는 것이 없는 날의 프로그램은 두 열을 다 쓴다
    예배 = next(c for c in sheet["cells"] if c["text"].startswith("봉사자 예배"))
    assert 예배["colspan"] == 2


# ---------------------------------------------------------------- 7. 빈 칸 병합


def test_07_내용_없는_칸도_병합돼_격자가_균일하다(sheet_data):
    """안 그러면 격자가 잘게 쪼개져 보인다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    empties = [c for c in sheet["cells"] if c["kind"] == "empty"]
    assert empties
    # 대부분은 2줄×2열 덩어리다
    full = [c for c in empties if c["rowspan"] == 2 and c["colspan"] == 2]
    assert len(full) > len(empties) / 2

    # 표에 그려지지 않는 자리가 하나도 없다
    covered = set()
    for cell in sheet["cells"]:
        for r in range(cell["row"], cell["row"] + cell["rowspan"]):
            for c in range(cell["col"], cell["col"] + cell["colspan"]):
                covered.add((r, c))
    missing = [(r, c) for r in range(sheet["total_rows"])
               for c in range(sheet["total_cols"]) if (r, c) not in covered]
    assert missing == []


# ---------------------------------------------------------------- 8. 색


def test_08_색이_정해진_대로_들어간다(sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)

    def fill_of(prefix):
        return next(c["fill"] for c in sheet["cells"]
                    if c["text"].startswith(prefix))

    assert fill_of("봉사자 예배") == "D9EAD3"      # 전체일정 내용
    assert fill_of("리허설") == "FFF2CC"           # 봉사자 내용 (공통)
    assert fill_of("음향 장비 확인") == "CFE2F3"   # 헤브론
    assert fill_of("무대 세팅") == "FCE5CD"        # 코람데오

    # **섞인 칸은 어느 팀의 것도 아니다** — 프로그램과 코람데오 항목이 같은
    # 시각이라 한 칸에 합쳐졌다. 첫째 것의 색을 쓰면 순서에 따라 색이 흔들린다.
    mixed = next(c for c in sheet["cells"] if c["text"].startswith("짐정리"))
    assert "짐 나르기" in mixed["text"]
    assert mixed["fill"] == staff_sheet.FILL_STAFF

    heads = [c for c in sheet["cells"] if c["kind"] == "head"]
    assert any(c["fill"] == "D9D9D9" for c in heads)          # 일자 머리
    assert any(c["text"] == "봉사자" and c["fill"] == "FFF2CC" for c in heads)
    assert any(c["text"] == "전체일정" and c["fill"] == "D9EAD3" for c in heads)
    assert all(c["fill"] is None for c in sheet["cells"] if c["kind"] == "empty")


def test_08b_봉사팀_파트_색은_파트_순서대로_배정된다():
    """목록을 코드에 박지 않는다 — 파트가 늘면 색도 늘어야 한다."""
    order = staff_sheet.team_parts()
    assert order == ["헤브론", "코람데오"]           # PROGRAM_PARTS 에 적힌 순서
    assert staff_sheet.fill_for_part(order[0]) == staff_sheet.TEAM_PART_FILLS[0]
    assert staff_sheet.fill_for_part(order[1]) == staff_sheet.TEAM_PART_FILLS[1]
    # 모르는 파트는 봉사자 공통색으로 떨어진다
    assert staff_sheet.fill_for_part("행정") == staff_sheet.FILL_STAFF


# ---------------------------------------------------------------- 9·10. 함정


def test_09_자정_넘긴_시각이_맨_아래에_온다(sheet_data):
    """`00:00 광고` 가 `06:00` 위에 서면 하루가 통째로 뒤집힌다."""
    assert staff_sheet.row_of("06:00") == 0
    assert staff_sheet.row_of("08:00") == 4
    assert staff_sheet.row_of("23:00") == 34
    assert staff_sheet.row_of("00:00") == 36        # 06:00 보다 아래
    assert staff_sheet.row_of("01:00") == 38
    assert staff_sheet.row_of("00:00") > staff_sheet.row_of("23:00")

    with app_session() as db:
        sheet = build(db, sheet_data)
    광고 = next(c for c in sheet["cells"] if "광고" in c["text"])
    아침 = next(c for c in sheet["cells"] if "아침식사" in c["text"])
    assert 광고["row"] > 아침["row"], "자정 넘긴 것이 위로 올라왔다"


def test_10_같은_줄에서_시작하는_프로그램이_한_칸에_합쳐진다(sheet_data):
    """그냥 넣으면 나중 것이 앞엣것을 덮어쓴다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    midnight = [c for c in sheet["cells"] if "광고" in c["text"]]
    assert len(midnight) == 1, "같은 시각인데 칸이 둘로 갈렸다"
    assert "광고" in midnight[0]["text"] and "야식" in midnight[0]["text"]
    assert "\n" in midnight[0]["text"]              # 줄바꿈으로 잇는다


# ---------------------------------------------------------------- 11·12. 내려받기


def test_11_내려받기를_로그인한_사람_누구나_할_수_있다(admin_client, sheet_data, client):
    response = admin_client.get("/live/staff.xlsx")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "봉사자시간표" in response.headers["content-disposition"] \
        or "%EB%B4%89" in response.headers["content-disposition"]

    assert client.get("/live/staff.xlsx", follow_redirects=False).status_code in (303, 401)


def test_12_화면의_표와_파일의_표가_같다(admin_client, sheet_data):
    """**같은 함수에서 나온 것**을 확인한다 — 따로 만들면 반드시 어긋난다."""
    import inspect

    # xlsx 쪽은 표를 만들지 않는다. 놓기만 한다.
    source = inspect.getsource(staff_xlsx)
    assert "staff_sheet.check_merges" in source
    assert "def build" not in source, "xlsx 쪽에서 표를 다시 만든다"
    assert "row_of" not in source, "xlsx 쪽에서 줄을 다시 계산한다"

    # 화면과 파일이 같은 칸을 그린다
    with app_session() as db:
        sheet = build(db, sheet_data)
        blob = staff_xlsx.write(sheet)

    page = admin_client.get("/live/staff").text
    book = load_workbook(io.BytesIO(blob))["봉사자 시간표"]

    for cell in sheet["cells"]:
        if not cell["text"]:
            continue
        excel = book.cell(row=cell["row"] + 1, column=cell["col"] + 1).value
        assert excel == cell["text"], f"파일이 다르다: {cell['text']!r} vs {excel!r}"
        # 화면에도 같은 글이 있다 (줄바꿈은 표시 방식이 다르므로 첫 줄로 본다)
        assert cell["text"].split("\n")[0][:12] in page

    # 병합 수도 같다
    merged = len(book.merged_cells.ranges)
    expected = sum(1 for c in sheet["cells"]
                   if c["rowspan"] > 1 or c["colspan"] > 1)
    assert merged == expected


# ---------------------------------------------------------------- 13. 병합 검사


def test_13_병합이_겹치면_저장하지_않고_멈춘다():
    """openpyxl 은 겹쳐도 아무 말이 없다. 손상된 파일을 만들어 놓고
    성공했다고 하면 안 된다."""
    broken = {
        "retreat": "시험",
        "columns": [{"kind": "time"}, {"kind": "content"}],
        "cells": [
            {"row": 0, "col": 0, "rowspan": 2, "colspan": 2, "text": "가",
             "fill": None, "kind": "body", "align": "left"},
            {"row": 1, "col": 1, "rowspan": 2, "colspan": 1, "text": "나",
             "fill": None, "kind": "body", "align": "left"},
        ],
        "header_rows": 0, "day_edges": [],
    }
    problems = staff_sheet.check_merges(broken["cells"])
    assert problems and "겹칩니다" in problems[0]

    with pytest.raises(staff_xlsx.SheetBroken) as caught:
        staff_xlsx.write(broken)
    assert "겹쳐" in str(caught.value)
    assert "만들지 않았습니다" in str(caught.value)


def test_13b_실제_표에는_겹치는_병합이_없다(sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)
    assert staff_sheet.check_merges(sheet["cells"]) == []


# ---------------------------------------------------------------- 14·15. 두 번째 시트


def test_14_두_번째_시트에_내려받은_시각과_사본_안내가_있다(sheet_data):
    """언제 것인지 파일이 스스로 말하지 않으면 그것이 곧 틀린 정보가 된다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    blob = staff_xlsx.write(sheet, now=dt.datetime(2026, 9, 1, 2, 30))

    book = load_workbook(io.BytesIO(blob))
    assert book.sheetnames == ["봉사자 시간표", "이 파일에 대해"]
    about = book["이 파일에 대해"]
    body = "\n".join(
        str(v) for row in about.iter_rows(values_only=True) for v in row if v
    )
    assert "2026 여름수련회 Belong" in body
    assert "2026-09-01 02:30" in body                # 벽시계 시각
    assert "내려받은 시점의 사본입니다" in body
    assert "바뀌면 다시 내려받으세요" in body
    assert "보는 법" in body


def test_14b_내려받은_시각은_벽시계다():
    """UTC 로 적으면 새벽에 하루 전이 된다 (첨부파일에서 겪은 그 문제)."""
    stamp = staff_sheet.now_label(dt.datetime(2026, 9, 1, 0, 30))
    assert stamp == "2026-09-01 00:30"
    assert staff_sheet.now_label().startswith(
        dt.datetime.now().strftime("%Y-%m-%d"))


def test_15_체크_상태가_파일에_없다(admin_client, sheet_data):
    """이 파일은 계획표지 현황판이 아니다."""
    with app_session() as db:
        item = db.scalars(select(models.ProgramItem).where(
            models.ProgramItem.text == "음향 장비 확인")).one()
        item.done_at = dt.datetime(2026, 8, 20, 11, 0)
        item.done_by_id = None
        db.commit()

    with app_session() as db:
        sheet = build(db, sheet_data)
    blob = staff_xlsx.write(sheet)
    book = load_workbook(io.BytesIO(blob))

    # 시간표 시트의 **내용 칸만** 본다 — 시각 열의 `11:00` 이나 안내 시트의
    # "체크 상태는 담지 않습니다" 는 체크 상태가 아니다
    body_cells = [c for c in sheet["cells"] if c["kind"] == "body"]
    body = "\n".join(c["text"] for c in body_cells)

    assert "음향 장비 확인" in body                  # 항목은 있고
    for mark in ("완료", "체크", "✓", "11:00", "08-20"):
        assert mark not in body, f"{mark} 가 표에 들어갔다"

    # 파일도 같다 — 화면과 파일이 같은 칸을 그리므로 한 번 더 확인한다
    for cell in body_cells:
        excel = book["봉사자 시간표"].cell(
            row=cell["row"] + 1, column=cell["col"] + 1).value
        assert excel == (cell["text"] or None)


# ---------------------------------------------------------------- 16. 빈 회차


def test_16_프로그램이_없는_회차에서도_빈_표가_정상적으로_나온다(admin_client):
    with app_session() as db:
        empty = models.Retreat(
            name="빈 회차", start_date=OPEN, end_date=CLOSE)
        db.add(empty)
        db.commit()
        sheet = staff_sheet.build(db, empty)

    # 일자는 있고 봉사자 칸만 있다 (전체일정은 없다)
    assert [d["name"] for d in sheet["days"]] == ["선발대", "1일차", "2일차", "폐회"]
    assert all(d["bands"] == ["staff"] for d in sheet["days"])
    assert sheet["cells"]
    assert staff_sheet.check_merges(sheet["cells"]) == []

    blob = staff_xlsx.write(sheet)
    assert len(blob) > 0
    book = load_workbook(io.BytesIO(blob))
    assert "봉사자 시간표" in book.sheetnames

    page = admin_client.get(f"/live/staff?retreat_id={empty.id}")
    assert page.status_code == 200


def test_16b_개회일이_없는_회차도_죽지_않는다(admin_client):
    with app_session() as db:
        blank = models.Retreat(name="개회일 미정")
        db.add(blank)
        db.commit()
        sheet = staff_sheet.build(db, blank)
    assert sheet["days"] == []
    assert staff_sheet.check_merges(sheet["cells"]) == []
    assert staff_xlsx.write(sheet)


# ════════════════════════════════════════════════════════════════════
#  다듬기 (5-8) — 화면으로 보고 걸린 것들
# ════════════════════════════════════════════════════════════════════


def body_texts(sheet):
    return [c["text"] for c in sheet["cells"] if c["kind"] == "body"]


def cell_with(sheet, needle):
    return next(c for c in sheet["cells"] if needle in c["text"])


# ---------------------------------------------------------------- 1·2. 겹친 말


def test_t01_같은_말이_한_칸에_두_번_나오지_않는다(sheet_data):
    """프로그램과 그 프로그램에 붙은 봉사팀 항목이 같은 말이면 하나만 낸다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    for name in ("골든벨 준비", "음향 및 무대설치", "홀 튜닝", "하차"):
        holders = [t for t in body_texts(sheet) if name in t]
        assert len(holders) == 1, f"{name} 이 여러 칸에 있다: {holders}"
        assert holders[0].count(name) == 1, f"한 칸에 두 번 나온다: {holders[0]!r}"


def test_t02_괄호_정보가_더_있는_쪽이_남는다(sheet_data):
    """`음향 및 무대설치` 와 `음향 및 무대설치 (3.5h)` 중 뒤엣것."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    assert cell_with(sheet, "음향 및 무대설치")["text"] == "음향 및 무대설치 (3.5h) (헤브론)"
    assert cell_with(sheet, "홀 튜닝")["text"] == "홀 튜닝 (1.5h) (헤브론)"

    # 괄호 정보가 없으면 프로그램 쪽 — 원래 시간표의 이름이다
    assert cell_with(sheet, "골든벨 준비")["text"] == "골든벨 준비 (담당M · 헤브론)"


def test_t02b_같고_다름은_이름이_서로_안에_들어_있는지로_본다():
    assert staff_sheet.same_thing("음향 및 무대설치", "음향 및 무대설치 (3.5h)")
    assert staff_sheet.same_thing("본당 도착 · 하차", "하차")
    assert staff_sheet.same_thing("복귀 짐 패킹 · 상차 · 이동",
                                  "헤브론 복귀 짐 패킹·상차·이동")
    # 구분자와 띄어쓰기가 달라도 같은 말로 본다
    assert staff_sheet.same_thing("짐 정리", "짐정리")
    # 다른 일은 합치지 않는다
    assert not staff_sheet.same_thing("짐정리", "짐 나르기")
    assert not staff_sheet.same_thing("저녁식사", "헤브론 집합 (17:00)")
    assert not staff_sheet.same_thing("", "무엇이든")


def test_t02c_버리는_쪽의_담당을_잃지_않는다(sheet_data):
    """`하차 (헤브론)` 를 버리면서 헤브론이 사라지면 안 된다 —
    이 표를 보는 사람이 정확히 그 정보를 보러 온다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    cell = cell_with(sheet, "본당 도착 · 하차")
    assert cell["text"] == "본당 도착 · 하차 (헤브론)"
    assert cell["fill"] == "CFE2F3", "합치고 나서 헤브론 색을 잃었다"


def test_t02d_다른_칸에_있으면_합치지_않는다(sheet_data):
    """프로그램이 전체일정이면 항목과 다른 칸이라 나란히 찍히지 않는다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    everyone = cell_with(sheet, "봉사자 예배")
    staff_side = cell_with(sheet, "음향 장비 확인")
    assert everyone["col"] != staff_side["col"]


# ---------------------------------------------------------------- 3·4. 눈금


def test_t03_한_시각의_줄_높이가_모두_같다(sheet_data):
    """시간표는 세로가 시간이어야 읽힌다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    times = [c for c in sheet["cells"] if c["kind"] == "time"]
    assert {c["rowspan"] for c in times} == {staff_sheet.ROWS_PER_HOUR}
    assert len(times) == 2 * staff_sheet.BODY_ROWS // staff_sheet.ROWS_PER_HOUR

    # 시각 칸이 정확히 한 시각씩 이어 붙는다 — 사이가 뜨거나 겹치지 않는다
    lefts = sorted(c["row"] for c in times if c["col"] == 0)
    assert lefts == list(range(
        sheet["header_rows"], sheet["total_rows"], staff_sheet.ROWS_PER_HOUR))

    # 높이는 구조가 정한다. 글 양이 아니다
    assert sheet["row_height"] == staff_sheet.ROW_HEIGHT_PX


def test_t04_파일의_줄_높이도_전부_같다(sheet_data):
    """예전에는 글 양에서 계산해 내용이 많은 시각이 두세 배로 부풀었다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]

    heights = {
        page.row_dimensions[r].height
        for r in range(sheet["header_rows"] + 1, sheet["total_rows"] + 1)
    }
    assert heights == {staff_xlsx.row_height(sheet)}


def test_t04b_넘치는_글은_줄을_늘리지_않고_작게_쓴다():
    """줄 높이가 고정이므로 넘치는 것은 글자를 줄여 담는다."""
    short, long = "짐정리", "가" * 60
    assert not staff_sheet.dense_of(short, 2, 1)
    assert staff_sheet.dense_of(long, 1, 1)
    # 넓은 칸이면 같은 글도 들어간다
    assert staff_sheet.wrapped_lines(long, 2) < staff_sheet.wrapped_lines(long, 1)
    # 작게 쓰면 더 들어간다
    assert (staff_sheet.wrapped_lines(long, 1, dense=True)
            <= staff_sheet.wrapped_lines(long, 1))


def test_t04c_실제_자료에서_넘치는_칸이_없다(sheet_data):
    """작게 써도 안 들어가면 글이 잘린다 — 그건 조용한 정보 손실이다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    over = [
        c for c in sheet["cells"] if c["kind"] == "body"
        and staff_sheet.wrapped_lines(c["text"], c["colspan"], dense=c["dense"])
        * (staff_sheet.DENSE_LINE_PX if c["dense"] else staff_sheet.LINE_PX)
        > c["rowspan"] * staff_sheet.ROW_HEIGHT_PX
    ]
    assert over == [], [c["text"] for c in over]


# ---------------------------------------------------------------- 5. 같은 줄


def test_t05_본당_도착_하차와_하차가_같은_줄에_온다(sheet_data):
    """실제 자료에서 둘은 같은 프로그램의 프로그램명과 항목이었다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    holders = [c for c in sheet["cells"] if "하차" in c["text"]]
    assert len(holders) == 1, "두 줄로 갈렸다"
    assert holders[0]["row"] - sheet["header_rows"] == staff_sheet.row_of("16:00")


# ---------------------------------------------------------------- 6·7·8. 괄호


def test_t06_담당이_총무팀이나_전체면_괄호가_나오지_않는다(sheet_data):
    """봉사팀에게 보내는 표다. 거의 모든 칸에 붙으면 정작 사람 이름이 묻힌다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    assert cell_with(sheet, "아침식사")["text"] == "아침식사"
    for text in body_texts(sheet):
        assert "(총무팀)" not in text
        assert "(전체)" not in text


def test_t07_사람_이름과_봉사팀_이름은_그대로_나온다(sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)

    assert "한지훈 목사" in cell_with(sheet, "봉사자 예배")["text"]
    assert "(헤브론)" in cell_with(sheet, "홀 튜닝")["text"]


def test_t07b_섞여_있으면_쓸모없는_것만_뺀다():
    assert staff_sheet.who_of("담당M · 헤브론 · 총무팀") == "담당M · 헤브론"
    assert staff_sheet.who_of("총무팀") == ""
    assert staff_sheet.who_of("전체") == ""
    assert staff_sheet.who_of("헤브론") == "헤브론"
    assert staff_sheet.who_of(None) == ""
    # **뺄 것이 없으면 원문 그대로** — 구분자를 다시 짜맞추지 않는다
    assert staff_sheet.who_of("하람·나윤") == "하람·나윤"
    assert staff_sheet.who_of("한지훈 목사 · 최도현M") == "한지훈 목사 · 최도현M"


def test_t08_장소는_그대로_나온다(sheet_data):
    """장소는 봉사팀에게 쓸모가 있다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    text = cell_with(sheet, "폐회 예배 · 성찬")["text"]
    assert text == "폐회 예배 · 성찬 (한지훈 목사 · 최도현M · 강당)"


# ---------------------------------------------------------------- 9·10. 안내


def test_t09_개회일_전체일정의_빈_구간에_안내가_나온다(sheet_data):
    """참가자가 아직 없어서 비어 있는 것이 맞습니다. 그 큰 공백이
    '뭔가 빠진 것' 처럼 보이지 않게 한 줄 적어 줍니다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    notes = [c for c in sheet["cells"] if c["kind"] == "note"]
    assert len(notes) == 1, "안내가 없거나 여러 날에 붙었다"
    note = notes[0]
    assert note["text"] == staff_sheet.NOTE_BEFORE_ARRIVAL == "참가자 등록 전"
    assert note["band"] == "all" and note["day"] == "1일차"   # 개회일
    assert note["row"] == sheet["header_rows"]                # 맨 위부터
    assert note["rowspan"] == staff_sheet.row_of("10:00")     # 첫 전체일정 전까지
    assert note["colspan"] == 2

    # 파일에도 같이 들어간다
    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]
    assert page.cell(row=note["row"] + 1, column=note["col"] + 1).value \
        == staff_sheet.NOTE_BEFORE_ARRIVAL


def test_t10_안내_칸에_색이_없다(sheet_data):
    """내용이 있는 것처럼 보이면 안 됩니다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    note = next(c for c in sheet["cells"] if c["kind"] == "note")
    assert note["fill"] is None

    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]
    target = page.cell(row=note["row"] + 1, column=note["col"] + 1)
    assert target.fill.fill_type in (None, "none")
    assert target.font.color.rgb.endswith("9B9A97"), "옅은 글씨가 아니다"


def test_t10b_개회일이_아니면_안내를_붙이지_않는다(admin_client):
    """폐회일 아침이 비어 있는 것은 참가자가 없어서가 아닙니다."""
    with app_session() as db:
        empty = models.Retreat(name="빈 회차", start_date=OPEN, end_date=CLOSE)
        db.add(empty)
        db.commit()
        sheet = staff_sheet.build(db, empty)
    assert [c for c in sheet["cells"] if c["kind"] == "note"] == []


# ---------------------------------------------------------------- 11. 여전히 하나


def test_t11_다듬은_뒤에도_화면과_파일이_같다(admin_client, sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)
        blob = staff_xlsx.write(sheet)

    page = admin_client.get("/live/staff").text
    book = load_workbook(io.BytesIO(blob))["봉사자 시간표"]

    for cell in sheet["cells"]:
        if not cell["text"]:
            continue
        assert book.cell(row=cell["row"] + 1,
                         column=cell["col"] + 1).value == cell["text"]
        assert cell["text"].split("\n")[0][:12] in page

    # 화면도 같은 줄 높이를 쓴다
    assert f"--rowh: {sheet['row_height']}px" in page \
        or f"--rowh:{sheet['row_height']}px" in page


# ════════════════════════════════════════════════════════════════════
#  칸의 비율과 새 회차 (5-8 · 5-1)
# ════════════════════════════════════════════════════════════════════


def test_r01_한_시각의_높이가_한_칸_너비의_2_8배쯤이다():
    """원본 시트에서 읽은 값 그대로. 세로가 시간인 표라 비율이 뜻을 가진다."""
    assert staff_sheet.SRC_ROW_PT == 28.5
    assert staff_sheet.SRC_CONTENT_COL == pytest.approx(20.06)
    assert staff_sheet.SRC_HOUR_PT == 57.0
    assert staff_sheet.HOUR_TO_WIDTH == pytest.approx(2.84, abs=0.01)


def test_r02_화면과_파일의_비율이_같다(sheet_data):
    """px 와 엑셀 단위로 단위는 다르지만 **같은 비율**이어야 한다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]

    # 파일 — 원본 값 그대로
    widths = [page.column_dimensions[get_column_letter(i + 1)].width
              for i in range(sheet["total_cols"])]
    assert widths[0] == staff_sheet.SRC_TIME_COL
    assert widths[-1] == staff_sheet.SRC_TIME_COL
    assert widths[1:3] == [staff_sheet.SRC_COL_MAIN, staff_sheet.SRC_COL_SIDE]
    assert page.row_dimensions[sheet["total_rows"]].height == staff_sheet.SRC_ROW_PT

    # 화면 — 같은 값을 px 로 옮긴 것
    screen = (sheet["row_height"] * staff_sheet.ROWS_PER_HOUR) / staff_sheet.COL_WIDTH_PX
    excel = (staff_sheet.SRC_HOUR_PT * staff_sheet.PT_TO_PX) / (
        staff_sheet.COL_MAIN_PX + staff_sheet.COL_SIDE_PX)
    assert screen == pytest.approx(excel, abs=0.02)


def test_r03_비율이_상수_하나에서_나온다():
    """**두 곳에서 따로 정하면 반드시 어긋난다** (5-8).

    xlsx 쪽에 열 너비나 줄 높이를 적어 두면 화면과 갈린다. 소스를 읽어
    그런 상수가 없는지 본다 — 규칙을 글로만 적어 두면 다음에 또 생긴다.
    """
    import inspect

    source = inspect.getsource(staff_xlsx)
    for banned in ("COL_WIDTH = ", "TIME_COL_WIDTH", "CONTENT_COL_WIDTH",
                   "BASE_ROW_HEIGHT"):
        assert banned not in source, f"xlsx 쪽에 {banned} 가 다시 생겼다"
    # 줄 높이도 구조에서 받아 쓴다
    assert "sheet.get(\"row_height_pt\"" in source or "row_height_pt" in source

    # 화면도 CSS 가 아니라 구조에서 너비를 받는다
    css = open("app/static/css/retreat.css", encoding="utf-8").read()
    at = css.index("봉사팀 보기 (CLAUDE.md 5-8)")
    assert "width:106px" not in css[at:]
    assert "width:54px" not in css[at:]


def test_r04_줄_높이는_여전히_어디서나_같다(sheet_data):
    """비율만 바꾸는 것이지 글 양에 따라 늘리는 것으로 되돌리지 않는다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]

    heights = {page.row_dimensions[r].height
               for r in range(sheet["header_rows"] + 1, sheet["total_rows"] + 1)}
    assert heights == {staff_sheet.SRC_ROW_PT}
    assert {c["rowspan"] for c in sheet["cells"] if c["kind"] == "time"} \
        == {staff_sheet.ROWS_PER_HOUR}


def test_r05_열_너비가_구조에_실려_화면으로_간다(admin_client, sheet_data):
    with app_session() as db:
        sheet = build(db, sheet_data)

    for column in sheet["columns"]:
        assert column["width"] > 0 and column["width_px"] > 0
    assert sheet["columns"][0]["width_px"] == staff_sheet.TIME_COL_PX

    page = admin_client.get("/live/staff").text
    assert f"width:{staff_sheet.COL_MAIN_PX}px" in page
    assert f"width:{staff_sheet.TIME_COL_PX}px" in page


# ---------------------------------------------------------------- 10. 새 회차


def test_r10_새_회차에_프로그램을_넣으면_표가_그려진다(admin_client):
    """**만들어 놓고 되겠지 하지 않는다.** 빈 회차에 다섯 가지를 넣고
    각각 제자리에 서는지 본다 — 겨울 회차에 실제로 있을 모양이다."""
    with app_session() as db:
        winter = models.Retreat(
            name="2027 겨울수련회", start_date=dt.date(2027, 1, 8),
            end_date=dt.date(2027, 1, 10))
        db.add(winter)
        db.flush()
        for order, (key, name) in enumerate(
                [("hebron", "5 헤브론"), ("koram", "6 코람데오")]):
            db.add(models.Department(retreat_id=winter.id, key=key, name=name,
                                     color_tag="#888888", sort_order=order))
        db.commit()
        winter_id = winter.id

    def post(body):
        res = admin_client.post(f"/live/program?retreat_id={winter_id}", json=body)
        assert res.status_code == 200, res.text
        return res.json()["id"]

    # 1. 봉사자만 하는 것
    post({"day": "1일차", "start_time": "08:00", "name": "아침식사",
          "host": "총무팀", "audience": "staff"})
    # 2. 정규일정
    rally = post({"day": "1일차", "start_time": "19:00", "name": "집회1",
                  "host": "오세영 목사"})
    # 3. 총무팀 작업 — 표에 안 나와야 한다
    post({"day": "1일차", "start_time": "17:00", "name": "강당 세팅",
          "host": "총무팀", "audience": "staff", "track": "ops"})
    # 4. 나란히 도는 것 — 오른쪽 열에 서야 한다
    post({"day": "1일차", "start_time": "19:00", "name": "새친구",
          "host": "하윤M", "end_time": "22:00", "parallel": True})
    # 5. 봉사팀 항목 — 봉사자 칸에 파랑으로
    res = admin_client.post(
        f"/live/program/{rally}/item?retreat_id={winter_id}",
        json={"phase": "pre", "part_key": "헤브론", "assignee_name": "헤브론",
              "text": "헤브론 집합", "scope": "team"})
    assert res.status_code == 200, res.text

    with app_session() as db:
        sheet = staff_sheet.build(db, db.get(models.Retreat, winter_id))

    def find(needle):
        return next((c for c in sheet["cells"] if needle in c["text"]), None)

    at19 = staff_sheet.row_of("19:00") + sheet["header_rows"]
    bands = {c["band"]: i for i, c in enumerate(sheet["columns"])
             if c.get("day") == "1일차" and c.get("half") == 0}

    # 1. 봉사자 칸
    breakfast = find("아침식사")
    assert breakfast and breakfast["col"] == bands["staff"]

    # 2. 전체일정 칸, 왼쪽 반 칸 (오른쪽에 나란히 도는 것이 있으므로)
    rally_cell = find("집회1")
    assert rally_cell["col"] == bands["all"] and rally_cell["colspan"] == 1
    assert rally_cell["row"] == at19

    # 3. **표에 없다**
    assert find("강당 세팅") is None

    # 4. 오른쪽 열
    guest = find("새친구")
    assert guest["col"] == bands["all"] + 1
    assert guest["row"] == at19

    # 5. 봉사자 칸에 헤브론 색
    gather = find("헤브론 집합")
    assert gather["col"] == bands["staff"]
    assert gather["row"] == at19
    assert gather["fill"] == staff_sheet.fill_for_part("헤브론")

    # 파일도 만들어진다
    assert staff_sheet.check_merges(sheet["cells"]) == []
    assert staff_xlsx.write(sheet)

    # 화면도 뜬다
    page = admin_client.get(f"/live/staff?retreat_id={winter_id}")
    assert page.status_code == 200
    assert "집회1" in page.text and "강당 세팅" not in page.text


# ════════════════════════════════════════════════════════════════════
#  글자 크기 (5-8) — 칸 크기와 같이 정한다
# ════════════════════════════════════════════════════════════════════


def test_f01_글자가_커졌고_화면과_파일이_같은_비율이다(sheet_data):
    """칸만 키우면 여백만 넓어진다. 글자도 줄 높이에서 끌어온다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    font = sheet["font"]

    # 화면의 글자/줄 = 파일의 글자/줄
    on_screen = font["body_px"] / sheet["row_height"]
    in_file = font["body_pt"] / sheet["row_height_pt"]
    assert on_screen == pytest.approx(in_file, abs=0.01)
    assert on_screen == pytest.approx(staff_sheet.FONT_TO_ROW, abs=0.01)

    # 원본(10pt : 28.5pt ≈ 0.35)보다 크다 — 여러 줄에 걸친 칸이 많아서다
    assert on_screen > 10 / 28.5

    # 예전 값보다 확실히 커졌다
    assert font["body_px"] > 11.5
    assert font["body_pt"] > 10


def test_f02_글자_크기가_비율_상수와_같은_자리에서_나온다():
    """**칸 크기와 글자 크기는 같이 정해야 한다** — 하나만 바꾸면 어긋난다."""
    assert staff_sheet.BODY_FONT_PT == pytest.approx(
        staff_sheet.SRC_ROW_PT * staff_sheet.FONT_TO_ROW, abs=0.05)
    assert staff_sheet.BODY_FONT_PX == round(
        staff_sheet.BODY_FONT_PT * staff_sheet.PT_TO_PX)

    import inspect

    # xlsx 쪽에 크기를 다시 적지 않았다
    source = inspect.getsource(staff_xlsx)
    for banned in ("FONT_SIZE = ", "DENSE_FONT_SIZE = ", "size=10", "size=8"):
        assert banned not in source, f"xlsx 쪽에 {banned} 가 생겼다"

    # CSS 쪽에도 적지 않았다 — 표에 실려 온 변수를 쓴다.
    # 표를 그리는 구획만 본다(다른 화면의 크기까지 묶으면 엉뚱한 것을 잡는다)
    css = open("app/static/css/retreat.css", encoding="utf-8").read()
    # `.sheetwrap` 부터가 표 자체다 — 그 위의 탭 줄은 화면 껍데기라 뺀다
    start = css.index(".sheetwrap{")
    end = css.index("프로그램 만들기·고치기 창", start)
    block = css[start:end]

    assert "font-size:var(--fs)" in block
    assert "font-size:var(--fs-dense)" in block
    assert "font-size:var(--fs-head)" in block
    for piece in block.split("font-size:")[1:]:
        value = piece.split(";")[0].split("}")[0].strip()
        assert value.startswith("var(--fs"), f"CSS 에 크기를 박았다: {value}"


def test_f03_dense_도_함께_커졌다(sheet_data):
    """기본이 커진 만큼 작은 글씨도 커져야 한다 — 안 그러면 지금처럼
    기본 대비 너무 작아진다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    font = sheet["font"]

    assert font["dense_px"] == round(
        staff_sheet.BODY_FONT_PT * staff_sheet.DENSE_FONT_SCALE
        * staff_sheet.PT_TO_PX)
    # 예전 값(화면 10px · 파일 8pt)보다 크다
    assert font["dense_px"] > 10
    assert font["dense_pt"] > 8
    # 그래도 본문보다는 작다
    assert font["dense_px"] < font["body_px"]
    assert font["dense_pt"] < font["body_pt"]


def test_f04_줄_높이는_글자를_키워도_그대로다(sheet_data):
    """**넘치면 넘치게 두고 줄을 늘리지 않는다** — 한 시각의 높이가
    어디서나 같아야 시각 눈금이 맞는다 (5-8)."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    assert sheet["row_height"] == staff_sheet.ROW_HEIGHT_PX
    assert sheet["row_height_pt"] == staff_sheet.SRC_ROW_PT

    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]
    heights = {page.row_dimensions[r].height
               for r in range(sheet["header_rows"] + 1, sheet["total_rows"] + 1)}
    assert heights == {staff_sheet.SRC_ROW_PT}


def test_f05_dense_가_걸리는_칸이_절반을_넘지_않는다(sheet_data):
    """절반이 넘으면 기본이 너무 큰 것이다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    body = [c for c in sheet["cells"] if c["kind"] == "body"]
    dense = [c for c in body if c["dense"]]
    assert body, "내용 칸이 하나도 없다"
    assert len(dense) * 2 < len(body), (
        f"{len(body)}칸 중 {len(dense)}칸이 작은 글씨 — 기본이 너무 크다")


def test_f06_머리줄과_시각_열이_본문보다_크고_굵다(sheet_data):
    """시각은 이 표에서 가장 자주 보는 글자다."""
    with app_session() as db:
        sheet = build(db, sheet_data)
    font = sheet["font"]
    assert font["head_px"] > font["body_px"]
    assert font["head_pt"] > font["body_pt"]

    blob = staff_xlsx.write(sheet)
    page = load_workbook(io.BytesIO(blob))["봉사자 시간표"]
    head = page.cell(row=1, column=1)                       # 왼쪽 위 '시각'
    time = page.cell(row=sheet["header_rows"] + 1, column=1)  # 06:00
    body = next(c for c in sheet["cells"]
                if c["kind"] == "body" and not c["dense"])
    plain = page.cell(row=body["row"] + 1, column=body["col"] + 1)

    assert head.font.size == font["head_pt"] and head.font.bold
    assert time.font.size == font["head_pt"] and time.font.bold
    assert plain.font.size == font["body_pt"] and not plain.font.bold

    # 화면도 같다
    css = open("app/static/css/retreat.css", encoding="utf-8").read()
    assert ".c-head > .cx,.c-time > .cx{font-size:var(--fs-head);font-weight:600}" in css


def test_f07_글자를_키운_만큼_줄_수도_다시_센다():
    """글자 폭을 크기에서 계산하지 않으면, 키운 뒤에도 예전 폭으로 세어
    넘치는 칸을 못 잡는다."""
    text = "음향 및 무대설치 (3.5h) (헤브론)"
    big = staff_sheet.wrapped_lines(text, 1)
    small = staff_sheet.wrapped_lines(text, 1, dense=True)
    assert small < big, "작게 써도 줄 수가 안 줄었다 — 폭을 크기에서 안 세고 있다"


def test_f08_실제_자료에서_잘리는_칸이_없다(sheet_data):
    """dense 로도 안 들어가면 글이 잘린다 — 조용한 정보 손실이다."""
    with app_session() as db:
        sheet = build(db, sheet_data)

    over = []
    for cell in sheet["cells"]:
        if cell["kind"] != "body":
            continue
        lines = staff_sheet.wrapped_lines(
            cell["text"], cell["colspan"], col=cell["col"], dense=cell["dense"])
        line_px = (staff_sheet.DENSE_LINE_PX if cell["dense"]
                   else staff_sheet.LINE_PX)
        if lines * line_px > cell["rowspan"] * staff_sheet.ROW_HEIGHT_PX:
            over.append(cell["text"])
    assert over == [], over
