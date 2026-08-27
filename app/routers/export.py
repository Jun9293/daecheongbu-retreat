"""엑셀 내보내기 (범용 포맷 — 정식 결산 양식은 Phase 3)."""

from __future__ import annotations

import io
import urllib.parse

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db import get_db
from app.deps import get_current_retreat
from app.domain.budget import build_budget_summary
from app.models import ExpenseEntry, Retreat, User
from app.security import get_current_user

router = APIRouter(prefix="/export")

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY = "#,##0"


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


@router.get("/expenses.xlsx")
def export_expenses(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    retreat: Retreat = Depends(get_current_retreat),
):
    entries = list(
        db.scalars(
            select(ExpenseEntry)
            .where(ExpenseEntry.retreat_id == retreat.id)
            .order_by(ExpenseEntry.receipt_number, ExpenseEntry.id)
        )
    )

    wb = Workbook()

    # 1) 지출 상세내역 — 기존 시트 컬럼 순서를 그대로 따른다
    ws = wb.active
    ws.title = "지출 상세내역"
    headers = [
        "구분",
        "항목",
        "세부항목-1",
        "세부항목-2",
        "세부항목-3",
        "부서",
        "영수증번호",
        "지출일자",
        "금액",
        "지원금액",
        "개인부담액",
        "식사인원",
        "참석자 명단",
        "비고",
        "지급여부",
        "지급일",
        "지출자",
        "지출자 계좌",
    ]
    ws.append(headers)
    for e in entries:
        ws.append(
            [
                e.level1,
                e.level2,
                e.level3a,
                e.level3b,
                e.level3c,
                e.department.name if e.department else None,
                e.receipt_number,
                e.expense_date,
                e.amount,
                e.subsidy_amount if e.is_meal_expense else e.amount,
                e.personal_burden_amount if e.is_meal_expense else 0,
                e.meal_headcount,
                " ".join(e.meal_attendee_names or []) or None,
                e.note,
                "지급완료" if e.paid else "미지급",
                e.paid_date,
                e.payer_name,
                e.payer_account,
            ]
        )
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=11):
        for cell in row:
            cell.number_format = MONEY
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    _style_header(ws, len(headers))
    _autosize(ws, [12, 16, 14, 14, 14, 10, 10, 12, 12, 12, 12, 9, 30, 24, 10, 12, 10, 20])

    # 2) 예산 대비 집행
    summary = build_budget_summary(db, retreat=retreat)
    ws2 = wb.create_sheet("예산 대비 집행")
    ws2.append(["구분", "항목", "세부항목", "예산금액", "집행금액", "잔액", "집행률(%)"])
    for row in summary.categories:
        cat = row.category
        ws2.append(
            [cat.level1, cat.level2, cat.level3, row.planned, row.spent, row.remaining, row.progress_pct]
        )
    ws2.append([])
    ws2.append(
        [
            "합계",
            "",
            "",
            summary.total_planned,
            summary.total_spent,
            summary.total_remaining,
            summary.progress_pct,
        ]
    )
    if summary.uncategorized_spent:
        ws2.append(["(예산 항목 미지정 지출)", "", "", 0, summary.uncategorized_spent, "", ""])
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = MONEY
    _style_header(ws2, 7)
    _autosize(ws2, [14, 20, 18, 14, 14, 14, 12])

    # 3) 환급 대상자
    ws3 = wb.create_sheet("환급 대상자")
    ws3.append(["지출자", "계좌번호", "환급 합계", "지급여부", "건수"])
    groups: dict[str, dict] = {}
    for e in entries:
        if e.subsidy_amount <= 0:
            continue
        key = f"{e.payer_name or '(미지정)'}|{e.payer_account or ''}"
        row = groups.setdefault(
            key,
            {
                "name": e.payer_name or "(미지정)",
                "account": e.payer_account or "",
                "total": 0,
                "count": 0,
                "unpaid": 0,
            },
        )
        row["total"] += e.subsidy_amount
        row["count"] += 1
        if not e.paid:
            row["unpaid"] += 1
    for row in sorted(groups.values(), key=lambda r: -r["total"]):
        ws3.append(
            [
                row["name"],
                row["account"],
                row["total"],
                "미지급 " + str(row["unpaid"]) + "건" if row["unpaid"] else "지급완료",
                row["count"],
            ]
        )
    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = MONEY
    _style_header(ws3, 5)
    _autosize(ws3, [14, 26, 14, 14, 8])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{retreat.name}_지출내역.xlsx"
    quoted = urllib.parse.quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted}"},
    )
