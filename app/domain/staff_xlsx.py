"""봉사자 시간표를 xlsx 로 쓴다 (CLAUDE.md 5-8).

**표를 여기서 만들지 않는다.** `staff_sheet.build()` 가 만든 것을 놓기만 한다 —
판정·색·병합을 여기서 다시 계산하면 화면과 어긋난다.

**글꼴은 맑은 고딕 10 이다.** 원본 시트는 Arial 이지만 한글 글자가 없어 대체
글꼴로 떨어지고, 그러면 열 너비에 맞춰 잡아 둔 자간이 어긋나 글이 잘린다.
**원본 재현보다 읽히는 것이 먼저다.**
"""

from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domain import staff_sheet

FONT = "맑은 고딕"

THIN = Side(style="thin", color="000000")
DOUBLE = Side(style="double", color="000000")

# **열 너비도 줄 높이도 글자 크기도 여기서 정하지 않는다.** 원본 시트에서
# 읽은 값이 `staff_sheet` 에 있고, 화면과 이 파일이 그 한 곳에서 끌어온다 —
# 두 곳에서 각각 정하면 화면과 파일이 어긋나고, 어긋난 쪽이 파일이면 아무도
# 모른 채 돌아다닌다 (5-8). **칸 크기와 글자 크기는 같이 정해야 한다.**


class SheetBroken(Exception):
    """병합이 겹쳤다. **저장하지 않고 멈춘다** — 손상된 파일을 만들어 놓고
    성공했다고 하면 안 된다."""


def _fill(color: str | None) -> PatternFill | None:
    if not color:
        return None
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def row_height(sheet: dict) -> float:
    """줄 높이. **어느 줄이나 같다.**

    예전에는 글 양에서 계산해 그 줄들이 나눠 갖게 했다. 그러면 내용이 많은
    시각이 두세 배로 부풀어 왼쪽 시각 눈금이 오른쪽 내용과 어긋난다 —
    시간표는 세로가 시간이어야 읽힌다. 넘치는 글은 줄을 늘리는 대신
    `dense` 로 작게 쓴다(그 판단은 `staff_sheet` 가 한다).
    """
    return sheet.get("row_height_pt", staff_sheet.SRC_ROW_PT)


def font_sizes(sheet: dict) -> dict:
    """글자 크기. **구조가 정한 것을 그대로 쓴다** — 칸 크기와 같은 자리에서
    나오므로 하나만 바뀌어 어긋나는 일이 없다 (5-8)."""
    return sheet.get("font") or {
        "body_pt": staff_sheet.BODY_FONT_PT,
        "head_pt": staff_sheet.HEAD_FONT_PT,
        "dense_pt": staff_sheet.DENSE_FONT_PT,
    }


def write(sheet: dict, *, now: dt.datetime | None = None) -> bytes:
    """구조를 xlsx 로. 돌려주는 것은 바이트다."""
    problems = staff_sheet.check_merges(sheet["cells"])
    if problems:
        # openpyxl 은 겹쳐도 아무 말이 없다. 우리가 말한다.
        raise SheetBroken(
            "병합이 겹쳐 파일을 만들지 않았습니다:\n  " + "\n  ".join(problems[:5])
        )

    sizes = font_sizes(sheet)
    book = Workbook()
    page = book.active
    page.title = "봉사자 시간표"
    page.sheet_view.showGridLines = False       # 격자선 표시를 끈다

    day_edges = set(sheet.get("day_edges") or [])

    for cell in sheet["cells"]:
        row, col = cell["row"] + 1, cell["col"] + 1
        target = page.cell(row=row, column=col)
        target.value = cell.get("text") or None
        big = cell["kind"] in ("head", "time")
        target.font = Font(
            name=FONT,
            size=(
                sizes["head_pt"] if big
                else sizes["dense_pt"] if cell.get("dense")
                else sizes["body_pt"]
            ),
            # 머리줄과 시각 열은 본문보다 크고 굵게 — 시각은 이 표에서
            # 가장 자주 보는 글자다
            bold=big,
            # 안내 글은 내용이 아니라 "여기는 비어 있는 것이 맞다" 는 말이라
            # 옅게 둔다 — 화면에서 흐린 글씨로 두는 것과 같은 뜻이다
            color="9B9A97" if cell["kind"] == "note" else "000000",
            italic=cell["kind"] == "note",
        )
        target.alignment = Alignment(
            horizontal=cell.get("align", "left"),
            vertical="top" if cell["kind"] == "body" else "center",
            wrap_text=True,
        )
        paint = _fill(cell.get("fill"))
        if paint:
            target.fill = paint

        if cell["rowspan"] > 1 or cell["colspan"] > 1:
            page.merge_cells(
                start_row=row, start_column=col,
                end_row=row + cell["rowspan"] - 1,
                end_column=col + cell["colspan"] - 1,
            )

        # **병합 덩어리는 바깥선만 그린다.** 안쪽에 줄이 비치면 지저분하다.
        # 일자 사이만 double.
        for r in range(row, row + cell["rowspan"]):
            for c in range(col, col + cell["colspan"]):
                left = THIN if c == col else None
                right = THIN if c == col + cell["colspan"] - 1 else None
                top = THIN if r == row else None
                bottom = THIN if r == row + cell["rowspan"] - 1 else None
                if left and (c - 1) in day_edges:
                    left = DOUBLE
                if right and c in day_edges:
                    right = DOUBLE
                page.cell(row=r, column=c).border = Border(
                    left=left, right=right, top=top, bottom=bottom
                )

    for index, column in enumerate(sheet["columns"], start=1):
        page.column_dimensions[get_column_letter(index)].width = column["width"]
    height = row_height(sheet)
    for row in range(sheet["header_rows"], sheet["total_rows"]):
        page.row_dimensions[row + 1].height = height
    for row in range(sheet["header_rows"]):
        page.row_dimensions[row + 1].height = sizes["head_pt"] * 1.7
    page.freeze_panes = page.cell(row=sheet["header_rows"] + 1, column=2)

    _about(book, sheet, now=now)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _about(book: Workbook, sheet: dict, *, now: dt.datetime | None) -> None:
    """두 번째 시트 — **언제 것인지 파일이 스스로 말하게 한다.**

    파일은 한 번 나가면 손을 떠나 카카오톡으로 돌아다니고, 사흘 뒤에도 누군가
    그걸 보고 있다. 언제 것인지 파일이 말하지 않으면 그것이 곧 틀린 정보가 된다.
    """
    page = book.create_sheet("이 파일에 대해")
    page.sheet_view.showGridLines = False
    lines = [
        ("회차", sheet["retreat"]),
        ("내려받은 시각", staff_sheet.now_label(now)),
        ("", ""),
        ("보는 법", "세로가 시각(06:00 ~ 다음날 01:00), 가로가 일자입니다."),
        ("", "일자마다 왼쪽이 봉사자 칸, 오른쪽이 전체일정 칸입니다."),
        ("", "전체일정이 없는 날(선발대)은 봉사자 칸만 있습니다."),
        ("", "시각 열은 양쪽 끝에 하나씩 있습니다 — 오른쪽에서도 볼 수 있게."),
        ("", ""),
        ("꼭 알아두세요",
         "이 파일은 내려받은 시점의 사본입니다. "
         "시스템에서 바꾼 것은 여기 반영되지 않습니다. 바뀌면 다시 내려받으세요."),
        ("", ""),
        ("담긴 것", "봉사자 일정과 봉사팀(헤브론·코람데오) 항목, 그리고 전체일정입니다."),
        ("", "총무팀 개인 항목과 뒤에서 도는 일(ops)은 담겨 있지 않습니다."),
        ("", "체크 상태는 담지 않습니다 — 이 파일은 계획표지 현황판이 아닙니다."),
    ]
    for index, (label, text) in enumerate(lines, start=1):
        left = page.cell(row=index, column=1, value=label or None)
        left.font = Font(name=FONT, size=font_sizes(sheet)["body_pt"], bold=True)
        left.alignment = Alignment(vertical="top")
        right = page.cell(row=index, column=2, value=text or None)
        right.font = Font(name=FONT, size=font_sizes(sheet)["body_pt"])
        right.alignment = Alignment(vertical="top", wrap_text=True)
    page.column_dimensions["A"].width = 16
    page.column_dimensions["B"].width = 74
